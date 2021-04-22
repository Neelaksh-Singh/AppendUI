from openpyxl import Workbook, load_workbook

# wb = load_workbook('try.xlsx')
# s = wb.active
# data = [("yash", 2)]

# for i in data:
#     s.append(i)
# wb.save('test.xlsx')
# print(data)


def appendSheet(content):
    wb = load_workbook('start.xlsx')
    sheet = wb.active
    for i in content:
        sheet.append(i)

    wb.save('final.xlsx')
